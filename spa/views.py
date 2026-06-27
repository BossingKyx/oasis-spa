"""Views for Oasis on the Go Spa — Phase 1 + customer self-booking."""
from datetime import datetime
from decimal import Decimal
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views.decorators.http import require_POST

from .availability import build_slots, is_slot_open
from .forms import BookingForm, CustomerForm, ExpenseForm, PaymentForm
from .models import (Booking, Branch, Customer, Expense, Payment, Service,
                     StaffProfile)


# ---------------------------------------------------------------------------
# Role-based access helpers
# ---------------------------------------------------------------------------
def get_profile(user):
    return getattr(user, 'profile', None)


def is_owner(user):
    p = get_profile(user)
    return bool(user.is_superuser or (p and p.is_owner))


def owner_required(view):
    """Block therapists from owner-only pages (sales, reports, expenses)."""
    @wraps(view)
    @login_required
    def _wrapped(request, *args, **kwargs):
        if not is_owner(request.user):
            return HttpResponseForbidden('This page is for the owner/admin only.')
        return view(request, *args, **kwargs)
    return _wrapped


def visible_bookings(user):
    """Owners see all bookings; therapists see only their own."""
    qs = Booking.objects.select_related('customer', 'branch', 'therapist__user')
    if is_owner(user):
        return qs
    p = get_profile(user)
    return qs.filter(therapist=p) if p else qs.none()


# ---------------------------------------------------------------------------
# Date / branch filter helpers
# ---------------------------------------------------------------------------
def parse_date(value, default):
    if value:
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            pass
    return default


def selected_branch(request):
    bid = request.GET.get('branch')
    if bid:
        return Branch.objects.filter(pk=bid).first()
    return None


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@login_required
def dashboard(request):
    user = request.user
    today = timezone.localdate()

    if not is_owner(user):
        # Therapist landing: their day at a glance.
        mine = visible_bookings(user).filter(scheduled_for__date=today)
        return render(request, 'spa/dashboard_staff.html', {
            'today': today,
            'my_bookings': mine.order_by('scheduled_for'),
            'active_count': mine.exclude(
                status__in=[Booking.CLOSED, Booking.CANCELLED, Booking.NO_SHOW]).count(),
        })

    start = parse_date(request.GET.get('from'), today.replace(day=1))
    end = parse_date(request.GET.get('to'), today)
    branch = selected_branch(request)

    payments = Payment.objects.filter(paid_at__date__gte=start, paid_at__date__lte=end)
    expenses = Expense.objects.filter(spent_on__gte=start, spent_on__lte=end)
    bookings = Booking.objects.filter(scheduled_for__date__gte=start,
                                      scheduled_for__date__lte=end)
    if branch:
        payments = payments.filter(booking__branch=branch)
        expenses = expenses.filter(branch=branch)
        bookings = bookings.filter(branch=branch)

    sales_total = payments.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    expense_total = expenses.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    top_services = (Service.objects.filter(bookings__in=bookings)
                    .annotate(n=Count('bookings'))
                    .order_by('-n')[:5])

    by_branch = []
    for b in Branch.objects.filter(is_active=True):
        bsales = payments.filter(booking__branch=b).aggregate(t=Sum('amount'))['t'] or 0
        by_branch.append({'branch': b, 'sales': bsales})

    pending_online = Booking.objects.filter(
        status=Booking.REQUESTED, external_source='self-booking').count()

    return render(request, 'spa/dashboard.html', {
        'start': start, 'end': end, 'branch': branch,
        'branches': Branch.objects.filter(is_active=True),
        'sales_total': sales_total,
        'expense_total': expense_total,
        'net': sales_total - expense_total,
        'booking_count': bookings.count(),
        'top_services': top_services,
        'by_branch': by_branch,
        'pending_online': pending_online,
    })


# ---------------------------------------------------------------------------
# Kanban service-flow board
# ---------------------------------------------------------------------------
@login_required
def board(request):
    user = request.user
    qs = visible_bookings(user).prefetch_related('services')

    date_val = parse_date(request.GET.get('date'), timezone.localdate())
    qs = qs.filter(scheduled_for__date=date_val)

    branch = selected_branch(request)
    if branch:
        qs = qs.filter(branch=branch)

    columns = []
    for status in Booking.BOARD_STATUSES:
        cards = [b for b in qs if b.status == status]
        columns.append({
            'key': status,
            'label': dict(Booking.STATUS_CHOICES)[status],
            'cards': cards,
        })

    return render(request, 'spa/board.html', {
        'columns': columns,
        'date': date_val,
        'branch': branch,
        'branches': Branch.objects.filter(is_active=True),
        'statuses': Booking.STATUS_CHOICES,
    })


@require_POST
@login_required
def booking_set_status(request, pk):
    """Move a card between columns; stamp start/finish times automatically."""
    booking = get_object_or_404(visible_bookings(request.user), pk=pk)
    new_status = request.POST.get('status')
    valid = dict(Booking.STATUS_CHOICES)
    if new_status not in valid:
        return JsonResponse({'ok': False, 'error': 'bad status'}, status=400)

    now = timezone.now()
    if new_status == Booking.IN_SERVICE and not booking.started_at:
        booking.started_at = now
    if new_status in (Booking.COMPLETED, Booking.PAID, Booking.CLOSED):
        if booking.started_at and not booking.finished_at:
            booking.finished_at = now
    booking.status = new_status
    booking.save()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        lt = timezone.localtime
        return JsonResponse({
            'ok': True,
            'status': booking.status,
            'started_at': lt(booking.started_at).strftime('%I:%M %p') if booking.started_at else '',
            'finished_at': lt(booking.finished_at).strftime('%I:%M %p') if booking.finished_at else '',
            'duration': booking.duration_label,
        })
    messages.success(request, f'Booking #{booking.pk} moved to {valid[new_status]}.')
    return redirect(request.META.get('HTTP_REFERER', 'board'))


# ---------------------------------------------------------------------------
# Bookings
# ---------------------------------------------------------------------------
@login_required
def booking_list(request):
    qs = visible_bookings(request.user).prefetch_related('services')
    status = request.GET.get('status')
    if status:
        qs = qs.filter(status=status)
    return render(request, 'spa/booking_list.html', {
        'bookings': qs[:200],
        'statuses': Booking.STATUS_CHOICES,
        'current_status': status,
    })


@login_required
def booking_detail(request, pk):
    booking = get_object_or_404(visible_bookings(request.user), pk=pk)
    return render(request, 'spa/booking_detail.html', {
        'booking': booking,
        'statuses': Booking.STATUS_CHOICES,
        'can_pay': is_owner(request.user),
    })


@owner_required
def booking_create(request):
    if request.method == 'POST':
        form = BookingForm(request.POST)
        if form.is_valid():
            booking = form.save(commit=False)
            booking.created_by = request.user
            booking.save()
            form.save_m2m()
            messages.success(request, f'Booking #{booking.pk} created.')
            return redirect('booking_detail', pk=booking.pk)
    else:
        initial = {'scheduled_for': timezone.localtime().strftime('%Y-%m-%dT%H:%M')}
        form = BookingForm(initial=initial)
    return render(request, 'spa/booking_form.html',
                  {'form': form, 'customer_form': CustomerForm()})


@owner_required
@require_POST
def customer_quick_create(request):
    """Add a customer inline from the booking form."""
    form = CustomerForm(request.POST)
    if form.is_valid():
        c = form.save()
        messages.success(request, f'Customer "{c.full_name}" added.')
    else:
        messages.error(request, 'Could not add customer — check the fields.')
    return redirect(request.META.get('HTTP_REFERER', 'booking_create'))


# ---------------------------------------------------------------------------
# Payments
# ---------------------------------------------------------------------------
@owner_required
def payment_create(request, pk):
    booking = get_object_or_404(Booking, pk=pk)
    if request.method == 'POST':
        form = PaymentForm(request.POST, request.FILES)
        if form.is_valid():
            pay = form.save(commit=False)
            pay.booking = booking
            pay.recorded_by = request.user
            pay.save()
            # Auto-advance to Paid when fully settled.
            if booking.is_paid and booking.status in (
                    Booking.COMPLETED, Booking.IN_SERVICE, Booking.ARRIVED):
                booking.status = Booking.PAID
                booking.save()
            messages.success(request, f'Payment of ₱{pay.amount:,.2f} recorded.')
            return redirect('booking_detail', pk=booking.pk)
    else:
        form = PaymentForm(initial={
            'amount': booking.balance,
            'paid_at': timezone.localtime().strftime('%Y-%m-%dT%H:%M'),
        })
    return render(request, 'spa/payment_form.html', {'form': form, 'booking': booking})


@login_required
def receipt(request, pk):
    booking = get_object_or_404(Booking, pk=pk)
    if not is_owner(request.user) and get_profile(request.user) != booking.therapist:
        return HttpResponseForbidden()
    return render(request, 'spa/receipt.html', {
        'booking': booking,
        'payments': booking.payments.all(),
        'now': timezone.localtime(),
    })


# ---------------------------------------------------------------------------
# Expenses / petty cash
# ---------------------------------------------------------------------------
@owner_required
def expense_list(request):
    today = timezone.localdate()
    start = parse_date(request.GET.get('from'), today.replace(day=1))
    end = parse_date(request.GET.get('to'), today)
    branch = selected_branch(request)
    qs = Expense.objects.filter(spent_on__gte=start, spent_on__lte=end)
    if branch:
        qs = qs.filter(branch=branch)
    total = qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    return render(request, 'spa/expense_list.html', {
        'expenses': qs.select_related('branch'),
        'total': total, 'start': start, 'end': end, 'branch': branch,
        'branches': Branch.objects.filter(is_active=True),
    })


@owner_required
def expense_create(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            exp = form.save(commit=False)
            exp.recorded_by = request.user
            exp.save()
            messages.success(request, f'Expense of ₱{exp.amount:,.2f} logged.')
            return redirect('expense_list')
    else:
        form = ExpenseForm(initial={'spent_on': timezone.localdate().strftime('%Y-%m-%d')})
    return render(request, 'spa/expense_form.html', {'form': form})


# ---------------------------------------------------------------------------
# Customers (CRM)
# ---------------------------------------------------------------------------
@login_required
def customer_list(request):
    q = request.GET.get('q', '').strip()
    qs = Customer.objects.all()
    if q:
        qs = qs.filter(Q(full_name__icontains=q) | Q(mobile__icontains=q) |
                       Q(facebook_name__icontains=q))
    return render(request, 'spa/customer_list.html', {'customers': qs[:200], 'q': q})


@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    return render(request, 'spa/customer_detail.html', {
        'customer': customer,
        'bookings': customer.bookings.select_related('branch').prefetch_related('services'),
    })


@owner_required
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            c = form.save()
            messages.success(request, f'Customer "{c.full_name}" added.')
            return redirect('customer_detail', pk=c.pk)
    else:
        form = CustomerForm()
    return render(request, 'spa/customer_form.html', {'form': form})


# ---------------------------------------------------------------------------
# Reports — daily transaction record + daily sales report (+ Excel/PDF export)
# ---------------------------------------------------------------------------
def _report_data(request):
    today = timezone.localdate()
    day = parse_date(request.GET.get('date'), today)
    branch = selected_branch(request)

    payments = (Payment.objects.filter(paid_at__date=day)
                .select_related('booking', 'booking__customer', 'booking__branch')
                .prefetch_related('booking__services'))
    if branch:
        payments = payments.filter(booking__branch=branch)

    rows = list(payments.order_by('paid_at'))
    sales_total = sum((p.amount for p in rows), Decimal('0'))

    by_method, by_branch, by_service = {}, {}, {}
    for p in rows:
        by_method[p.method] = by_method.get(p.method, Decimal('0')) + p.amount
        bname = p.booking.branch.name
        by_branch[bname] = by_branch.get(bname, Decimal('0')) + p.amount
    # Service-type counts come from the day's bookings.
    bookings = Booking.objects.filter(scheduled_for__date=day)
    if branch:
        bookings = bookings.filter(branch=branch)
    for bk in bookings.prefetch_related('services'):
        for s in bk.services.all():
            by_service[s.name] = by_service.get(s.name, 0) + 1

    expenses = Expense.objects.filter(spent_on=day)
    if branch:
        expenses = expenses.filter(branch=branch)
    expense_total = expenses.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    return {
        'day': day, 'branch': branch, 'rows': rows,
        'sales_total': sales_total, 'expense_total': expense_total,
        'net': sales_total - expense_total,
        'by_method': by_method, 'by_branch': by_branch, 'by_service': by_service,
        'expenses': list(expenses.select_related('branch')),
    }


@owner_required
def daily_report(request):
    data = _report_data(request)
    data['branches'] = Branch.objects.filter(is_active=True)
    return render(request, 'spa/daily_report.html', data)


@owner_required
def daily_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font
    data = _report_data(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Transactions'
    title = f"Daily Report - {data['day']:%b %d, %Y}"
    if data['branch']:
        title += f" - {data['branch'].name}"
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Time', 'Booking #', 'Customer', 'Branch', 'Services', 'Method', 'Amount']
    ws.append([])
    ws.append(headers)
    for c in ws[3]:
        c.font = Font(bold=True)
    for p in data['rows']:
        ws.append([
            timezone.localtime(p.paid_at).strftime('%I:%M %p'),
            p.booking.pk, p.booking.customer.full_name, p.booking.branch.name,
            p.booking.services_label, p.method, float(p.amount),
        ])
    ws.append([])
    ws.append(['', '', '', '', '', 'Sales total', float(data['sales_total'])])
    ws.append(['', '', '', '', '', 'Expenses', float(data['expense_total'])])
    ws.append(['', '', '', '', '', 'Net', float(data['net'])])

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = f"oasis_daily_{data['day']:%Y%m%d}.xlsx"
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


@owner_required
def daily_report_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    data = _report_data(request)

    resp = HttpResponse(content_type='application/pdf')
    fname = f"oasis_daily_{data['day']:%Y%m%d}.pdf"
    resp['Content-Disposition'] = f'inline; filename="{fname}"'
    c = canvas.Canvas(resp, pagesize=A4)
    w, h = A4
    y = h - 20 * mm
    c.setFont('Helvetica-Bold', 16)
    c.drawString(20 * mm, y, 'Oasis on the Go Spa')
    y -= 7 * mm
    c.setFont('Helvetica', 11)
    sub = f"Daily Report - {data['day']:%b %d, %Y}"
    if data['branch']:
        sub += f" - {data['branch'].name}"
    c.drawString(20 * mm, y, sub)
    y -= 10 * mm

    c.setFont('Helvetica-Bold', 9)
    cols = [20, 35, 70, 120, 150, 175]
    for label, x in zip(['Time', 'Customer', 'Services', 'Branch', 'Method', 'Amount'], cols):
        c.drawString(x * mm, y, label)
    y -= 2 * mm
    c.line(20 * mm, y, 195 * mm, y)
    y -= 5 * mm
    c.setFont('Helvetica', 8)
    for p in data['rows']:
        if y < 25 * mm:
            c.showPage()
            y = h - 20 * mm
            c.setFont('Helvetica', 8)
        c.drawString(20 * mm, y, timezone.localtime(p.paid_at).strftime('%I:%M%p'))
        c.drawString(35 * mm, y, p.booking.customer.full_name[:20])
        c.drawString(70 * mm, y, p.booking.services_label[:28])
        c.drawString(120 * mm, y, p.booking.branch.name[:14])
        c.drawString(150 * mm, y, p.method[:12])
        c.drawRightString(195 * mm, y, f"{p.amount:,.2f}")
        y -= 5 * mm

    y -= 3 * mm
    c.line(20 * mm, y, 195 * mm, y)
    y -= 6 * mm
    c.setFont('Helvetica-Bold', 10)
    c.drawRightString(175 * mm, y, 'Sales total:')
    c.drawRightString(195 * mm, y, f"{data['sales_total']:,.2f}")
    y -= 6 * mm
    c.drawRightString(175 * mm, y, 'Expenses:')
    c.drawRightString(195 * mm, y, f"{data['expense_total']:,.2f}")
    y -= 6 * mm
    c.setFillColor(colors.HexColor('#2e6b3e'))
    c.drawRightString(175 * mm, y, 'Net:')
    c.drawRightString(195 * mm, y, f"{data['net']:,.2f}")
    c.showPage()
    c.save()
    return resp


# ---------------------------------------------------------------------------
# Customer self-booking (public, no login)
# ---------------------------------------------------------------------------
def public_book(request):
    """Public page: pick branch + service(s) + date, see open slots, request."""
    branches = Branch.objects.filter(is_active=True)
    services = Service.objects.filter(is_active=True)
    today = timezone.localdate()

    def _selected(key, many=False):
        if request.method == 'POST':
            return request.POST.getlist(key) if many else request.POST.get(key, '')
        return request.GET.getlist(key) if many else request.GET.get(key, '')

    branch_id = _selected('branch')
    date_str = _selected('date')
    service_ids = _selected('services', many=True)

    branch = branches.filter(pk=branch_id).first() if branch_id else None
    sel_date = parse_date(date_str, None) if date_str else None
    sel_services = list(services.filter(pk__in=service_ids)) if service_ids else []

    slots = build_slots(branch, sel_date) if (branch and sel_date) else []
    errors = []

    if request.method == 'POST':
        # Honeypot: real users leave this blank.
        if request.POST.get('website'):
            return redirect('public_book')

        name = request.POST.get('full_name', '').strip()
        mobile = request.POST.get('mobile', '').strip()
        slot_raw = request.POST.get('slot', '')
        service_type = request.POST.get('service_type', Booking.WALK_IN)
        home_address = request.POST.get('home_address', '').strip()
        notes = request.POST.get('notes', '').strip()
        slot_dt = parse_datetime(slot_raw) if slot_raw else None

        if not branch:
            errors.append('Please choose a branch.')
        if not sel_services:
            errors.append('Please choose at least one service.')
        if not name:
            errors.append('Please enter your name.')
        if not mobile:
            errors.append('Please enter your mobile number.')
        if not slot_dt:
            errors.append('Please pick an available time slot.')
        if service_type == Booking.HOME and not home_address:
            errors.append('Please enter your address for home service.')
        if slot_dt and branch and not is_slot_open(branch, slot_dt):
            errors.append('Sorry, that time was just taken. Please pick another slot.')

        if not errors:
            customer = (Customer.objects.filter(mobile=mobile).first()
                        if mobile else None)
            if not customer:
                customer = Customer.objects.create(full_name=name, mobile=mobile)
            booking = Booking.objects.create(
                customer=customer, branch=branch, service_type=service_type,
                channel='Website', status=Booking.REQUESTED, scheduled_for=slot_dt,
                home_address=home_address, notes=notes,
                external_source='self-booking')
            booking.services.set(sel_services)
            return redirect('public_book_done', pk=booking.pk)

    return render(request, 'spa/public_book.html', {
        'branches': branches, 'services': services,
        'branch': branch, 'sel_date': sel_date, 'sel_services': sel_services,
        'sel_service_ids': [str(s.pk) for s in sel_services],
        'slots': slots, 'errors': errors, 'today': today,
        'min_date': today.strftime('%Y-%m-%d'),
        'posted': request.POST if request.method == 'POST' else {},
    })


def public_book_done(request, pk):
    """Public confirmation page — limited, non-sensitive details only."""
    booking = get_object_or_404(Booking, pk=pk, external_source='self-booking')
    return render(request, 'spa/public_done.html', {'booking': booking})


# ---------------------------------------------------------------------------
# AI receipt scanning — auto-fill expense fields from an uploaded photo
# ---------------------------------------------------------------------------
@owner_required
@require_POST
def expense_scan_receipt(request):
    """Read a receipt photo with Claude vision and return structured fields."""
    import base64
    import io
    import json
    import os
    import urllib.error
    import urllib.request

    from django.conf import settings as _s

    upload = request.FILES.get('image')
    if not upload:
        return JsonResponse({'ok': False, 'error': 'Please choose a receipt photo first.'})
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return JsonResponse({'ok': False, 'error':
            'AI auto-fill is not set up yet. Ask the admin to add an ANTHROPIC_API_KEY.'})

    # Normalize + downscale the image (controls cost and avoids media-type issues).
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(upload.read())).convert('RGB')
        img.thumbnail((1600, 1600))
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=85)
        b64 = base64.standard_b64encode(buf.getvalue()).decode()
    except Exception:
        return JsonResponse({'ok': False, 'error': 'That file did not look like an image.'})

    categories = ', '.join(_s.OASIS.get('EXPENSE_CATEGORIES', []))
    prompt = (
        "This is a photo of a Philippine sales receipt, official receipt, or invoice. "
        "Extract these fields. Use an empty string for anything not clearly visible — "
        "do not guess or invent values.\n"
        "- supplier_name: the store/business that issued the receipt\n"
        "- supplier_tin: the supplier's TIN (tax id), digits/dashes only\n"
        "- supplier_address: the supplier's address\n"
        "- reference: the receipt/invoice/OR number\n"
        "- amount: the TOTAL amount due, as a plain number with no currency symbol or commas\n"
        "- spent_on: the receipt date as YYYY-MM-DD, or empty if not shown\n"
        f"- category: the best fit from this list (or empty): {categories}"
    )
    schema = {
        "type": "object",
        "properties": {k: {"type": "string"} for k in (
            "supplier_name", "supplier_tin", "supplier_address",
            "reference", "amount", "spent_on", "category")},
        "required": ["supplier_name", "supplier_tin", "supplier_address",
                     "reference", "amount", "spent_on", "category"],
        "additionalProperties": False,
    }
    model = _s.OASIS.get('RECEIPT_MODEL', 'claude-opus-4-8')
    body = json.dumps({
        "model": model,
        "max_tokens": 1024,
        "messages": [{"role": "user", "content": [
            {"type": "image", "source": {
                "type": "base64", "media_type": "image/jpeg", "data": b64}},
            {"type": "text", "text": prompt},
        ]}],
        "output_config": {"format": {"type": "json_schema", "schema": schema}},
    }).encode()
    # Call the Anthropic Messages API directly (stdlib only — keeps the
    # serverless bundle small enough to deploy on Vercel).
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=body, method="POST",
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                 "content-type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            data = json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        detail = e.read().decode(errors='ignore')[:200]
        return JsonResponse({'ok': False, 'error': f'AI service error ({e.code}): {detail}'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'AI service error: {e}'})

    text = next((b.get('text', '') for b in data.get('content', [])
                 if b.get('type') == 'text'), '')
    try:
        fields = json.loads(text)
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error':
            'Could not read that receipt clearly. Please type the details in.'})
    return JsonResponse({'ok': True, 'fields': fields})
